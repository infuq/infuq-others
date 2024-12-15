#! /usr/bin/env python

import openpyxl
import datetime
import re


class WriteXlsxUtil(object):

    def __init__(self, file_name='File'):
        self.suffix = '.xlsx'
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.create_sheet(index=0, title="Sheet")
        self.file_name = file_name
        self.empty = True  # 标记是否真的写入了内容
        self.row = 0
    
    # 设置列宽
    def column_width(self, index, width):
        self.sheet.column_dimensions[index].width = width    

    # 写入单元格内容
    def write(self, content):
        self.empty = False
        self.row += 1
        for index in range(1, len(content) + 1):  # index = column
            if isinstance(content[index - 1], str) or isinstance(content[index - 1], (int, float)):
                v = str(content[index - 1])
                digit = re.match(r'^\d+(\.\d+)?$', v, re.I)
                if digit:
                    self.sheet.cell(self.row, index).value = v.zfill(len(v)).data_type = "int"
                else:
                    self.sheet.cell(self.row, index).value = v
            else:
                cell = content[index - 1]
                self.sheet.cell(self.row, index).value = cell['value']
                if 'color' in cell:
                    self.sheet.cell(self.row, index).font = openpyxl.styles.Font(color=cell['color'])

    # 保存文件
    def save(self):
        if not self.empty:
            self.workbook.save("./" + self.file_name + "-" + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + self.suffix)
