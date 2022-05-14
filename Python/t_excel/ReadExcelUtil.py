#! /usr/bin/env python

'''
读取xlsx或xls文件

pip install xlrd
pip install xlwt
pip install openpyxl
'''

import openpyxl
import os
import xlrd
from collections.abc import Iterable
from collections.abc import Iterator


class ReadExcelUtil(Iterable):

    def __init__(self, file_name):
        self.filename = file_name
        self.suffix = os.path.splitext(self.filename)[-1][1:].lower()# xls xlsx

        if self.suffix == 'xlsx':
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheetnames = self.workbook.sheetnames# Sheet1 Sheet2 Sheet3...
            self.sheet = self.workbook[self.sheetnames[0]]# 根据名称默认读取第一个Sheet

        if self.suffix == 'xls':
            self.workbook = xlrd.open_workbook(self.filename, encoding_override="UTF-8")
            self.sheetnames = self.workbook._sheet_names# Sheet1 Sheet2 Sheet3...
            # self.sheet = self.workbook[0]# 根据下标默认读取第一个Sheet
            self.sheet = self.workbook[self.sheetnames[0]]# 根据名称默认读取第一个Sheet

    def __iter__(self):
        if self.suffix == 'xlsx':
            return ReadXlsxIterator(self.sheet)
        if self.suffix == 'xls':
            return ReadXlsIterator(self.sheet)


# 迭代器 xlsx
class ReadXlsxIterator(Iterator):

    def __init__(self, sheet):
        self.sheet = sheet
        self.max_row = sheet.max_row
        self.max_column = sheet.max_column
        self.row = 0

    def __next__(self):
        self.row += 1
        try:
            if self.row <= self.max_row:
                # 读取行内容
                row_data = []
                for i in range(1, self.max_column + 1):
                    cell_value = self.sheet.cell(row=self.row, column=i).value
                    row_data.append(cell_value)
                # (行号, 行内容)
                return (self.row, row_data)
        except Exception:
            raise StopIteration
        raise StopIteration


# 迭代器 xls
class ReadXlsIterator(Iterator):

    def __init__(self, sheet):
        self.sheet = sheet
        self.nrows = sheet.nrows
        self.row = -1

    def __next__(self):
        self.row += 1
        try:
            if self.row <= self.nrows:
                # 读取行内容
                # (行号, 行内容)
                return (self.row, self.sheet.row_values(self.row))
        except Exception:
            raise StopIteration
        raise StopIteration

