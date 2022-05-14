#! /usr/bin/env python

import openpyxl
import datetime


class WriteXlsxUtil(object):

    def __init__(self, file_name='File'):
        self.suffix = '.xlsx'
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.create_sheet(index=0, title="Sheet")
        self.file_name = file_name
        self.empty = True  # 标记是否真的写入了内容
        self.row = 0

    # 写入单元格内容
    def write(self, content):
        self.empty = False
        self.row += 1
        for index in range(1, len(content) + 1):  # index = column
            v = str(content[index - 1])
            if v.isdigit():
                self.sheet.cell(self.row, index).value = v.zfill(len(v))
            else:
                self.sheet.cell(self.row, index).value = v

    # 保存文件
    def save(self):
        if not self.empty:
            self.workbook.save("./" + self.file_name + "-" + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + self.suffix)
